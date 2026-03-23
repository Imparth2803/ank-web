"""
core/exporter.py
----------------
Builds the formatted Excel output using openpyxl.

Column layout (10 cols total):

  Sale (DR offset):
    DATE | BILL NO. | DEBIT AMT. | TD | CD | UD | ANK AMT. | CREDIT AMT. | DAYS | ANK AMT.
    col:    1    2       3         4   5   6     7             8             9      10

  Purchase (CR offset):
    DATE | BILL NO. | DEBIT AMT. | DAYS | ANK AMT. | CREDIT AMT. | TD | CD | UD | ANK AMT.
    col:    1    2       3          4      5             6           7   8   9     10

  TD = Total Days  (raw days before offset)
  CD = Offset Days (user-entered, same every row)
  UD = Used Days   = TD - CD  <- used in ANK formula
  DAYS = raw days, no offset (single column, shown as "188 Days")
"""

from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.config import EXCEL_STYLE as S
from core.summarizer import Summary


@dataclass
class ExportMeta:
    company_name:    str
    client_name:     str
    from_date:       str
    to_date:         str
    interest_rate:   float
    debit_days:      int   = 0
    credit_days:     int   = 0
    manual_interest: float = 0.0
    ledger_type:     str   = "sale"
    output_filename: str   = ""


# ---------------------------------------------------------------------------
# Column maps (10 columns each)
#
# Sale:
#   1=DATE  2=BILL  3=DR_AMT  4=DR_TD  5=DR_CD  6=DR_UD  7=DR_ANK
#   8=CR_AMT  9=CR_DAYS  10=CR_ANK
#
# Purchase:
#   1=DATE  2=BILL  3=DR_AMT  4=DR_DAYS  5=DR_ANK
#   6=CR_AMT  7=CR_TD  8=CR_CD  9=CR_UD  10=CR_ANK
# ---------------------------------------------------------------------------

_SALE_COLS = dict(
    date=1, bill=2,
    dr_amt=3, dr_td=4, dr_cd=5, dr_ud=6, dr_ank=7,
    cr_amt=8, cr_days=9, cr_ank=10,
)

_PUR_COLS = dict(
    date=1, bill=2,
    dr_amt=3, dr_days=4, dr_ank=5,
    cr_amt=6, cr_td=7, cr_cd=8, cr_ud=9, cr_ank=10,
)


def _cols(ledger_type: str) -> dict:
    return _PUR_COLS if ledger_type == "purchase" else _SALE_COLS


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export(
    df: pd.DataFrame,
    summary: Summary,
    meta: ExportMeta,
    output_path: str | Path,
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "ANK Statement"

    cols = _cols(meta.ledger_type)

    _set_column_widths(ws, meta.ledger_type)
    row = 1
    row = _write_title(ws, meta, row)
    row = _write_meta_rows(ws, meta, row)
    row += 1
    row = _write_column_headers(ws, meta.ledger_type, row)
    data_start = row
    row = _write_data_rows(ws, df, meta, cols, row)
    data_end = row - 1
    row = _write_totals(ws, meta.ledger_type, cols, data_start, data_end, row)
    row += 1
    row = _write_summary(ws, summary, meta, cols, data_start, data_end, row)

    output_path = Path(output_path)
    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Section writers
# ---------------------------------------------------------------------------

def _write_title(ws, meta: ExportMeta, row: int) -> int:
    label = f"{'SALE' if meta.ledger_type != 'purchase' else 'PURCHASE'} — Company: {meta.company_name}"
    cell = ws.cell(row=row, column=1, value=label)
    cell.font      = Font(name=S.header_font_name, bold=True, size=13, color=S.accent_font)
    cell.fill      = PatternFill("solid", start_color=S.accent_fill)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    return row + 1


def _write_meta_rows(ws, meta: ExportMeta, row: int) -> int:
    is_sale      = meta.ledger_type != "purchase"
    offset_label = "DR DAYS:" if is_sale else "CR DAYS:"
    offset_value = str(meta.debit_days if is_sale else meta.credit_days)
    if offset_value == "0":
        offset_value = "-"

    pairs_r1 = [
        ("From Date:", meta.from_date),
        ("To Date:",   meta.to_date),
        ("INTEREST:",  f"{meta.interest_rate}%"),
        (offset_label, offset_value),
    ]
    col = 1
    for label, value in pairs_r1:
        _bold_cell(ws, row, col, label)
        _plain_cell(ws, row, col + 1, value)
        col += 2
    row += 1

    _bold_cell(ws, row, 1, "Client Name:")
    _plain_cell(ws, row, 2, meta.client_name)
    if meta.manual_interest:
        _bold_cell(ws, row, 7, "Manual Int.:")
        _plain_cell(ws, row, 8, f"{meta.manual_interest:.2f}")
    return row + 1


def _write_column_headers(ws, ledger_type: str, row: int) -> int:
    if ledger_type == "purchase":
        # DR side: single DAYS | CR side: TD CD UD
        headers = [
            "DATE", "BILL NO.",
            "DEBIT AMT.", "DAYS", "ANK AMT.",
            "CREDIT AMT.", "TD", "CD", "UD", "ANK AMT.",
        ]
    else:
        # DR side: TD CD UD | CR side: single DAYS
        headers = [
            "DATE", "BILL NO.",
            "DEBIT AMT.", "TD", "CD", "UD", "ANK AMT.",
            "CREDIT AMT.", "DAYS", "ANK AMT.",
        ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font      = Font(name=S.header_font_name, bold=True, color=S.header_font, size=10)
        cell.fill      = PatternFill("solid", start_color=S.header_fill)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = _thin_border()
    ws.row_dimensions[row].height = 18
    return row + 1


def _write_data_rows(ws, df: pd.DataFrame, meta: ExportMeta, cols: dict, row: int) -> int:
    is_sale  = meta.ledger_type != "purchase"
    cd_value = meta.debit_days if is_sale else meta.credit_days

    for _, r in df.iterrows():
        date_str = r["date"].strftime("%d-%m-%Y") if pd.notna(r["date"]) else ""
        ws.cell(row=row, column=cols["date"], value=date_str)
        ws.cell(row=row, column=cols["bill"], value=r["bill_no"])

        if is_sale:
            # Debit side — TD / CD / UD / ANK
            if r["debit"] > 0:
                ud = int(r["debit_days_col"])   # UD (already offset-subtracted)
                td = ud + cd_value              # TD = UD + CD
                _num_cell(ws, row, cols["dr_amt"], r["debit"],         color=S.debit_font)
                _num_cell(ws, row, cols["dr_td"],  td)
                _num_cell(ws, row, cols["dr_cd"],  cd_value)
                _num_cell(ws, row, cols["dr_ud"],  ud)
                _num_cell(ws, row, cols["dr_ank"], round(r["debit_ank"]))
            # Credit side — single DAYS / ANK
            if r["credit"] > 0:
                raw_days = int(r["credit_days_col"])  # no offset on credit for sale
                _num_cell(ws, row, cols["cr_amt"],  r["credit"],       color=S.credit_font)
                _plain_cell(ws, row, cols["cr_days"], f"{raw_days} Days")
                _num_cell(ws, row, cols["cr_ank"],  round(r["credit_ank"]))
        else:
            # Purchase
            # Debit side — single DAYS / ANK
            if r["debit"] > 0:
                raw_days = int(r["debit_days_col"])   # no offset on debit for purchase
                _num_cell(ws, row, cols["dr_amt"],  r["debit"],        color=S.debit_font)
                _plain_cell(ws, row, cols["dr_days"], f"{raw_days} Days")
                _num_cell(ws, row, cols["dr_ank"],  round(r["debit_ank"]))
            # Credit side — TD / CD / UD / ANK
            if r["credit"] > 0:
                ud = int(r["credit_days_col"])
                td = ud + cd_value
                _num_cell(ws, row, cols["cr_amt"], r["credit"],        color=S.credit_font)
                _num_cell(ws, row, cols["cr_td"],  td)
                _num_cell(ws, row, cols["cr_cd"],  cd_value)
                _num_cell(ws, row, cols["cr_ud"],  ud)
                _num_cell(ws, row, cols["cr_ank"], round(r["credit_ank"]))

        row += 1
    return row


def _write_totals(ws, ledger_type: str, cols: dict, data_start: int, data_end: int, row: int) -> int:
    fill = PatternFill("solid", start_color=S.total_fill)

    ws.cell(row=row, column=1, value="TOTAL:").font = Font(name=S.body_font_name, bold=True)
    ws.cell(row=row, column=1).fill = fill

    sum_cols = [
        (cols["dr_amt"], S.debit_font),
        (cols["dr_ank"], S.debit_font),
        (cols["cr_amt"], S.credit_font),
        (cols["cr_ank"], S.credit_font),
    ]

    for col_idx, color in sum_cols:
        letter = get_column_letter(col_idx)
        c = ws.cell(row=row, column=col_idx,
                    value=f"=SUM({letter}{data_start}:{letter}{data_end})")
        c.font          = Font(name=S.body_font_name, bold=True, color=color)
        c.fill          = fill
        c.number_format = "#,##0"

    for c in range(1, 11):
        cell = ws.cell(row=row, column=c)
        if cell.value is None:
            cell.fill = fill

    return row + 1


def _write_summary(
    ws, s: Summary, meta: ExportMeta, cols: dict,
    data_start: int, data_end: int, row: int
) -> int:
    tot_row  = row - 2
    rate     = meta.interest_rate / 100

    dr_l     = get_column_letter(cols["dr_amt"])
    cr_l     = get_column_letter(cols["cr_amt"])
    dr_ank_l = get_column_letter(cols["dr_ank"])
    cr_ank_l = get_column_letter(cols["cr_ank"])

    # Net Balance / Total ANK
    _bold_cell(ws, row, 1, "Net Balance:")
    ws.cell(row=row, column=2,
            value=f"={dr_l}{tot_row}-{cr_l}{tot_row}").number_format = "#,##0.00"
    _bold_cell(ws, row, 4, "Total ANK:")
    ws.cell(row=row, column=5,
            value=f"={dr_ank_l}{tot_row}-{cr_ank_l}{tot_row}").number_format = "#,##0.00"
    row += 2

    # Interest
    _bold_cell(ws, row, 1, "Interest")
    if meta.manual_interest:
        c = ws.cell(row=row, column=3, value=meta.manual_interest)
        c.font = Font(name=S.body_font_name, color="0000FF")
    else:
        c = ws.cell(row=row, column=3,
                    value=f"=({dr_l}{tot_row}*{rate})/30")
    c.number_format = "#,##0.00"
    interest_ref = f"C{row}"
    row += 1

    # Receivable
    _bold_cell(ws, row, 1, "Receivable:")
    c = ws.cell(row=row, column=3,
                value=f"=({dr_ank_l}{tot_row}-{cr_ank_l}{tot_row})*{rate}")
    c.number_format = "#,##0.00"
    receivable_ref = f"C{row}"
    row += 1

    # Average Days
    _bold_cell(ws, row, 1, "Average Days:")
    c = ws.cell(row=row, column=3,
                value=f"=IFERROR({receivable_ref}/{interest_ref},0)")
    c.number_format = "#,##0.00"
    c.font = Font(name=S.body_font_name, bold=True)

    is_sale = meta.ledger_type != "purchase"
    offset  = meta.debit_days if is_sale else meta.credit_days
    if offset:
        side = "DR" if is_sale else "CR"
        note = f"{side} offset: -{offset}d  (UD = TD - CD)"
        ws.cell(row=row, column=6, value=note).font = Font(
            name=S.body_font_name, italic=True, color="888888", size=9)

    return row + 1


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _set_column_widths(ws, ledger_type: str) -> None:
    if ledger_type == "purchase":
        # DATE | BILL | DR_AMT | DAYS | DR_ANK | CR_AMT | TD | CD | UD | CR_ANK
        widths = [13, 10, 13, 11, 13, 13, 10, 10, 10, 13]
    else:
        # DATE | BILL | DR_AMT | TD | CD | UD | DR_ANK | CR_AMT | DAYS | CR_ANK
        widths = [13, 10, 13, 10, 10, 10, 13, 13, 11, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(bottom=s)


def _bold_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=S.body_font_name, bold=True)
    return c


def _plain_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=S.body_font_name)
    return c


def _num_cell(ws, row, col, value, color: str = "000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font          = Font(name=S.body_font_name, color=color)
    c.number_format = "#,##0"
    c.alignment     = Alignment(horizontal="right")
    return c